import csv
import hashlib
import io
import uuid
from datetime import datetime, date
from typing import Optional
from sqlalchemy.orm import Session

from app.models import Entry, Category, Account
from app.schemas.external_source import (
    CSVColumnMapping,
    CSVPreviewRow,
    CSVUploadResponse,
    ImportConfirmResponse,
)

# In-memory storage for uploaded files (in production, use Redis or file storage)
_file_cache: dict[str, dict] = {}


def parse_csv(file_content: bytes, encoding: str = "utf-8") -> tuple[list[str], list[dict]]:
    """Parse CSV content and return headers and rows"""
    try:
        content = file_content.decode(encoding)
    except UnicodeDecodeError:
        # Try other encodings
        for enc in ["cp949", "euc-kr", "utf-8-sig"]:
            try:
                content = file_content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Unable to decode CSV file")

    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rows = list(reader)

    return headers, rows


def make_headers_unique(headers: list[str]) -> list[str]:
    """Make duplicate header names unique by appending _2, _3, etc."""
    seen = {}
    unique_headers = []
    for header in headers:
        if header in seen:
            seen[header] += 1
            unique_headers.append(f"{header}_{seen[header]}")
        else:
            seen[header] = 1
            unique_headers.append(header)
    return unique_headers


def parse_excel(file_content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Parse Excel file (.xls or .xlsx) and return headers and rows"""
    file_ext = filename.lower().split(".")[-1] if filename else ""

    if file_ext == "xls":
        # Use xlrd for .xls files
        try:
            import xlrd
        except ImportError:
            raise ValueError("xlrd library is required for .xls files")

        workbook = xlrd.open_workbook(file_contents=file_content)
        sheet = workbook.sheet_by_index(0)

        if sheet.nrows < 1:
            raise ValueError("Excel file is empty")

        # Get headers from first row
        raw_headers = []
        for col in range(sheet.ncols):
            cell_value = sheet.cell_value(0, col)
            raw_headers.append(str(cell_value).strip() if cell_value else f"Column{col+1}")
        headers = make_headers_unique(raw_headers)

        # Get data rows
        rows = []
        for row_idx in range(1, sheet.nrows):
            row_data = {}
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                cell_value = cell.value

                # Handle date cells
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                        cell_value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                    except:
                        pass
                # Handle number cells
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Keep as number but convert to string for consistency
                    if cell_value == int(cell_value):
                        cell_value = str(int(cell_value))
                    else:
                        cell_value = str(cell_value)
                else:
                    cell_value = str(cell_value).strip() if cell_value else ""

                if col_idx < len(headers):
                    row_data[headers[col_idx]] = cell_value

            # Skip empty rows
            if any(v for v in row_data.values()):
                rows.append(row_data)

        return headers, rows

    elif file_ext == "xlsx":
        # Use openpyxl for .xlsx files
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl library is required for .xlsx files")

        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active

        if sheet is None or sheet.max_row < 1:
            raise ValueError("Excel file is empty")

        # Get headers from first row
        raw_headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            raw_headers.append(str(cell_value).strip() if cell_value else f"Column{col}")
        headers = make_headers_unique(raw_headers)

        # Get data rows
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value

                # Handle date cells
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%Y-%m-%d")
                elif isinstance(cell_value, date):
                    cell_value = cell_value.strftime("%Y-%m-%d")
                elif cell_value is not None:
                    # Handle numbers
                    if isinstance(cell_value, float):
                        if cell_value == int(cell_value):
                            cell_value = str(int(cell_value))
                        else:
                            cell_value = str(cell_value)
                    else:
                        cell_value = str(cell_value).strip()
                else:
                    cell_value = ""

                if col_idx - 1 < len(headers):
                    row_data[headers[col_idx - 1]] = cell_value

            # Skip empty rows
            if any(v for v in row_data.values()):
                rows.append(row_data)

        workbook.close()
        return headers, rows

    else:
        raise ValueError(f"Unsupported Excel format: {file_ext}")


def parse_file(file_content: bytes, filename: str, encoding: str = "utf-8") -> tuple[list[str], list[dict]]:
    """Parse file based on extension and return headers and rows"""
    file_ext = filename.lower().split(".")[-1] if filename else ""

    if file_ext == "csv":
        return parse_csv(file_content, encoding)
    elif file_ext in ("xls", "xlsx"):
        return parse_excel(file_content, filename)
    else:
        raise ValueError(f"Unsupported file format: {file_ext}. Supported formats: csv, xls, xlsx")


def detect_column_mapping(headers: list[str]) -> CSVColumnMapping:
    """Detect column mapping based on header names"""
    mapping = CSVColumnMapping()

    # Common Korean column names
    date_names = ["날짜", "거래일", "일자", "date", "거래일시", "사용일"]
    amount_names = ["금액", "거래금액", "amount", "출금", "입금", "사용금액", "결제금액"]
    type_names = ["수입/지출", "유형", "거래유형", "type", "구분", "입출금구분"]
    category_names = ["카테고리", "분류", "category", "업종", "가맹점업종"]
    subcategory_names = ["소분류", "세부분류", "subcategory"]
    memo_names = ["메모", "비고", "적요", "memo", "내용", "거래내용", "사용처", "가맹점명", "가맹점"]
    account_names = ["계좌", "통장", "account", "카드", "자산"]

    for header in headers:
        header_lower = header.lower().strip()
        if any(name in header_lower for name in date_names):
            if not mapping.date:  # Only set if not already set
                mapping.date = header
        elif any(name in header_lower for name in amount_names):
            if not mapping.amount:
                mapping.amount = header
        elif any(name in header_lower for name in type_names):
            if not mapping.type:
                mapping.type = header
        elif any(name in header_lower for name in subcategory_names):
            if not mapping.subcategory:
                mapping.subcategory = header
        elif any(name in header_lower for name in category_names):
            if not mapping.category:
                mapping.category = header
        elif any(name in header_lower for name in memo_names):
            if not mapping.memo:
                mapping.memo = header
        elif any(name in header_lower for name in account_names):
            if not mapping.account:
                mapping.account = header

    return mapping


def parse_date(date_str: str) -> Optional[date]:
    """Parse date string to date object"""
    if not date_str or not isinstance(date_str, str):
        return None

    date_str = date_str.strip()

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y년 %m월 %d일",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y.%m.%d %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    return None


def parse_amount(amount_str: str) -> Optional[int]:
    """Parse amount string to integer"""
    if not amount_str:
        return None

    if isinstance(amount_str, (int, float)):
        return int(amount_str)

    # Remove currency symbols, commas, spaces
    cleaned = str(amount_str).replace(",", "").replace(" ", "").replace("원", "")
    cleaned = cleaned.replace("₩", "").replace("$", "").replace("\\", "")

    try:
        # Handle negative amounts
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        return int(float(cleaned))
    except ValueError:
        return None


def generate_row_hash(row: dict, columns: list[str]) -> str:
    """Generate a hash for duplicate detection"""
    values = [str(row.get(col, "")) for col in sorted(columns)]
    return hashlib.md5("|".join(values).encode()).hexdigest()


def preview_import(
    db: Session,
    file_content: bytes,
    household_id: uuid.UUID,
    filename: str = "file.csv",
    encoding: str = "utf-8",
) -> CSVUploadResponse:
    """Parse file and generate preview"""
    headers, rows = parse_file(file_content, filename, encoding)
    mapping = detect_column_mapping(headers)

    # Generate file ID and cache
    file_id = str(uuid.uuid4())
    _file_cache[file_id] = {
        "headers": headers,
        "rows": rows,
        "household_id": str(household_id),
        "created_at": datetime.utcnow(),
    }

    # Generate preview rows (first 10)
    preview_rows = []
    for i, row in enumerate(rows[:10]):
        date_str = row.get(mapping.date, "") if mapping.date else ""
        amount_str = row.get(mapping.amount, "") if mapping.amount else ""

        parsed_date = parse_date(date_str)
        parsed_amount = parse_amount(amount_str)

        error = None
        if not parsed_date:
            error = f"Invalid date: {date_str}"
        elif parsed_amount is None:
            error = f"Invalid amount: {amount_str}"

        # Determine type
        entry_type = "expense"
        if mapping.type and row.get(mapping.type):
            type_value = str(row.get(mapping.type, "")).lower()
            if "income" in type_value or "수입" in type_value or "입금" in type_value:
                entry_type = "income"
            elif "transfer" in type_value or "이체" in type_value:
                entry_type = "transfer"
        elif parsed_amount and parsed_amount > 0:
            # Positive might be income depending on bank format
            entry_type = "expense"  # Default to expense, user can adjust

        preview_rows.append(
            CSVPreviewRow(
                row_number=i + 1,
                date=str(date_str),
                amount=abs(parsed_amount) if parsed_amount else 0,
                type=entry_type,
                category=str(row.get(mapping.category, "")) if mapping.category and row.get(mapping.category) else None,
                memo=str(row.get(mapping.memo, "")) if mapping.memo and row.get(mapping.memo) else None,
                account=str(row.get(mapping.account, "")) if mapping.account and row.get(mapping.account) else None,
                is_duplicate=False,  # Will be checked during confirm
                error=error,
            )
        )

    return CSVUploadResponse(
        file_id=file_id,
        total_rows=len(rows),
        preview_rows=preview_rows,
        detected_columns=headers,
        suggested_mapping=mapping,
    )


def execute_import(
    db: Session,
    file_id: str,
    household_id: uuid.UUID,
    user_id: uuid.UUID,
    column_mapping: CSVColumnMapping,
    default_account_id: Optional[uuid.UUID],
    default_category_id: Optional[uuid.UUID],
    default_payer_member_id: uuid.UUID,
    skip_duplicates: bool = True,
) -> ImportConfirmResponse:
    """Execute the import from cached file data"""
    cached = _file_cache.get(file_id)
    if not cached:
        return ImportConfirmResponse(
            imported_count=0,
            skipped_count=0,
            error_count=0,
            errors=["File not found. Please upload again."],
        )

    if str(household_id) != cached["household_id"]:
        return ImportConfirmResponse(
            imported_count=0,
            skipped_count=0,
            error_count=0,
            errors=["Invalid file ID."],
        )

    rows = cached["rows"]
    imported_count = 0
    skipped_count = 0
    error_count = 0
    errors = []

    # Get existing entry hashes for duplicate detection
    existing_hashes = set()
    if skip_duplicates:
        existing_entries = db.query(Entry).filter(
            Entry.household_id == household_id
        ).all()
        for entry in existing_entries:
            hash_str = f"{entry.date}|{entry.amount}|{entry.memo or ''}"
            existing_hashes.add(hashlib.md5(hash_str.encode()).hexdigest())

    # Get categories for matching
    categories = db.query(Category).filter(
        (Category.household_id == None) | (Category.household_id == household_id)
    ).all()
    category_map = {c.name.lower(): c.id for c in categories}

    # Build category aliases for fuzzy matching
    category_aliases = {
        # 식비 관련
        "음식": "식비", "식사": "식비", "밥": "식비", "점심": "식비", "저녁": "식비",
        "아침": "식비", "배달": "식비", "외식": "식비", "식당": "식비",
        # 교통비 관련
        "교통": "교통비", "버스": "교통비", "지하철": "교통비", "택시": "교통비",
        "기차": "교통비", "ktx": "교통비", "고속버스": "교통비", "주유": "교통비",
        # 주거비 관련
        "주거": "주거비", "월세": "주거비", "관리비": "주거비", "전기세": "주거비",
        "수도세": "주거비", "가스비": "주거비", "인터넷": "주거비",
        # 통신비 관련
        "통신": "통신비", "휴대폰": "통신비", "핸드폰": "통신비", "전화": "통신비",
        # 의료비 관련
        "의료": "의료비", "병원": "의료비", "약국": "의료비", "약": "의료비", "건강": "의료비",
        # 문화/여가 관련
        "문화": "문화/여가", "여가": "문화/여가", "영화": "문화/여가", "공연": "문화/여가",
        "취미": "문화/여가", "운동": "문화/여가", "헬스": "문화/여가", "게임": "문화/여가",
        # 쇼핑 관련
        "쇼핑": "쇼핑", "의류": "쇼핑", "옷": "쇼핑", "마트": "쇼핑", "편의점": "쇼핑",
        # 교육 관련
        "교육": "교육비", "학원": "교육비", "책": "교육비", "강의": "교육비",
        # 급여 관련
        "급여": "급여", "월급": "급여", "보너스": "급여", "상여": "급여",
        # 기타
        "기타": "기타", "미분류": "기타",
    }

    def find_category_id(name: str) -> uuid.UUID | None:
        """Find category ID with fuzzy matching"""
        if not name:
            return None
        name_lower = name.lower().strip()

        # Exact match first
        if name_lower in category_map:
            return category_map[name_lower]

        # Try alias match
        if name_lower in category_aliases:
            alias_target = category_aliases[name_lower].lower()
            if alias_target in category_map:
                return category_map[alias_target]

        # Try partial match (category name contains the search term or vice versa)
        for cat_name, cat_id in category_map.items():
            if name_lower in cat_name or cat_name in name_lower:
                return cat_id

        return None

    # Get existing accounts for matching (and track created ones)
    existing_accounts = db.query(Account).filter(
        Account.household_id == household_id
    ).all()
    account_map = {a.name.lower(): a.id for a in existing_accounts}

    for i, row in enumerate(rows):
        try:
            date_str = row.get(column_mapping.date, "") if column_mapping.date else ""
            amount_str = row.get(column_mapping.amount, "") if column_mapping.amount else ""

            parsed_date = parse_date(date_str)
            parsed_amount = parse_amount(amount_str)

            if not parsed_date:
                error_count += 1
                errors.append(f"Row {i+1}: Invalid date '{date_str}'")
                continue

            if parsed_amount is None:
                error_count += 1
                errors.append(f"Row {i+1}: Invalid amount '{amount_str}'")
                continue

            # Determine type
            entry_type = "expense"
            if column_mapping.type and row.get(column_mapping.type):
                type_value = str(row.get(column_mapping.type, "")).lower()
                if "income" in type_value or "수입" in type_value or "입금" in type_value:
                    entry_type = "income"
                elif "transfer" in type_value or "이체" in type_value:
                    entry_type = "transfer"

            memo = str(row.get(column_mapping.memo, "")) if column_mapping.memo and row.get(column_mapping.memo) else None

            # Check for duplicates
            if skip_duplicates:
                hash_str = f"{parsed_date}|{abs(parsed_amount)}|{memo or ''}"
                row_hash = hashlib.md5(hash_str.encode()).hexdigest()
                if row_hash in existing_hashes:
                    skipped_count += 1
                    continue
                existing_hashes.add(row_hash)

            # Find category with fuzzy matching
            category_id = None
            if column_mapping.category and row.get(column_mapping.category):
                category_name = str(row.get(column_mapping.category, "")).strip()
                category_id = find_category_id(category_name)

            # Use default category as fallback if no match found
            if category_id is None and default_category_id:
                category_id = default_category_id

            # Get subcategory
            subcategory = None
            if column_mapping.subcategory and row.get(column_mapping.subcategory):
                subcategory = str(row.get(column_mapping.subcategory, "")).strip() or None

            # Find or create account
            account_id = default_account_id
            if column_mapping.account and row.get(column_mapping.account):
                account_name = str(row.get(column_mapping.account, "")).strip()
                if account_name:
                    account_name_lower = account_name.lower()
                    if account_name_lower in account_map:
                        account_id = account_map[account_name_lower]
                    else:
                        # Auto-create account
                        new_account = Account(
                            owner_user_id=user_id,
                            household_id=household_id,
                            name=account_name,
                            type="shared",
                            is_shared_visible=True,
                        )
                        db.add(new_account)
                        db.flush()  # Get the ID
                        account_map[account_name_lower] = new_account.id
                        account_id = new_account.id

            # Create entry
            entry = Entry(
                household_id=household_id,
                created_by_user_id=user_id,
                type=entry_type,
                amount=abs(parsed_amount),
                date=parsed_date,
                occurred_at=datetime.combine(parsed_date, datetime.min.time()),
                category_id=category_id,
                subcategory=subcategory,
                memo=memo,
                payer_member_id=default_payer_member_id,
                shared=False,
                account_id=account_id,
            )
            db.add(entry)
            imported_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {i+1}: {str(e)}")

    db.commit()

    # Clean up cache
    del _file_cache[file_id]

    return ImportConfirmResponse(
        imported_count=imported_count,
        skipped_count=skipped_count,
        error_count=error_count,
        errors=errors[:20],  # Limit error messages
    )


def check_duplicates(
    db: Session,
    household_id: uuid.UUID,
    entries: list[dict],
) -> list[bool]:
    """Check which entries are duplicates"""
    existing_entries = db.query(Entry).filter(
        Entry.household_id == household_id
    ).all()

    existing_hashes = set()
    for entry in existing_entries:
        hash_str = f"{entry.date}|{entry.amount}|{entry.memo or ''}"
        existing_hashes.add(hashlib.md5(hash_str.encode()).hexdigest())

    results = []
    for entry_data in entries:
        hash_str = f"{entry_data.get('date')}|{entry_data.get('amount')}|{entry_data.get('memo', '')}"
        row_hash = hashlib.md5(hash_str.encode()).hexdigest()
        results.append(row_hash in existing_hashes)

    return results
